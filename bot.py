import os
import pytz
import logging
import asyncio
import requests
from datetime import datetime, time, timedelta
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters,
    ContextTypes, ConversationHandler
)
import openpyxl
from openpyxl import Workbook
import re

# ✅ Устанавливаем часовой пояс
TIMEZONE = pytz.timezone('Europe/Moscow')

def get_current_datetime():
    return datetime.now(TIMEZONE)

def get_current_time():
    return get_current_datetime().time()

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Константы для состояний разговора
WAITING_TIME, WAITING_LUNCH_CONFIRMATION, WAITING_DESCRIPTION, WAITING_REMINDER_TIME = range(4)

# Импорт конфигурации
from config import BOT_TOKEN, EXCEL_FILE, DEFAULT_REMINDER_HOUR, DEFAULT_REMINDER_MINUTE, USER_SETTINGS, WELCOMED_USERS, MAX_ENTRIES_PER_DAY, YANDEX_DISK_ENABLED, YANDEX_DISK_TOKEN, YANDEX_DISK_FOLDER

# ✅ Глобальная ссылка на application для доступа к job_queue
global_app = None

class YandexDiskManager:
    def __init__(self, token: str):
        self.token = token
        self.base_url = "https://cloud-api.yandex.net/v1/disk/resources"
        self.headers = {
            "Authorization": f"OAuth {token}",
            "Content-Type": "application/json"
        }

    def create_folder(self, folder_path: str):
        """Создает папку на Яндекс.Диске"""
        try:
            url = f"{self.base_url}?path={folder_path}"
            response = requests.put(url, headers=self.headers)
            if response.status_code in [200, 201, 409]:  # 409 - уже существует
                print(f"✅ Папка на Яндекс.Диске создана или уже существует: {folder_path}")
                return True
            else:
                print(f"❌ Ошибка создания папки: {response.status_code} - {response.text}")
                return False
        except Exception as e:
            print(f"❌ Ошибка при создании папки: {e}")
            return False

    def upload_file(self, local_file_path: str, remote_file_path: str):
        """Загружает файл на Яндекс.Диск"""
        try:
            # Получаем URL для загрузки
            url = f"{self.base_url}/upload?path={remote_file_path}&overwrite=true"
            response = requests.get(url, headers=self.headers)
            
            if response.status_code != 200:
                print(f"❌ Ошибка получения URL для загрузки: {response.status_code} - {response.text}")
                return False
            
            upload_url = response.json()["href"]
            
            # Загружаем файл
            with open(local_file_path, 'rb') as file:
                upload_response = requests.put(upload_url, files={"file": file})
            
            if upload_response.status_code in [200, 201]:
                print(f"✅ Файл успешно загружен на Яндекс.Диск: {remote_file_path}")
                return True
            else:
                print(f"❌ Ошибка загрузки файла: {upload_response.status_code} - {upload_response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Ошибка при загрузке файла: {e}")
            return False

    def get_file_info(self, file_path: str):
        """Получает информацию о файле на Яндекс.Диске"""
        try:
            url = f"{self.base_url}?path={file_path}"
            response = requests.get(url, headers=self.headers)
            if response.status_code == 200:
                return response.json()
            else:
                return None
        except Exception as e:
            print(f"❌ Ошибка получения информации о файле: {e}")
            return None

# ✅ Инициализация менеджера Яндекс.Диска
yandex_disk = YandexDiskManager(YANDEX_DISK_TOKEN) if YANDEX_DISK_ENABLED and YANDEX_DISK_TOKEN else None

class ExcelManager:
    def __init__(self, filename: str):
        self.filename = filename
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        """Создаёт файл, если не существует. НЕ удаляем активный лист."""
        try:
            directory = os.path.dirname(self.filename)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
                print(f"✅ Создана папка: {directory}")

            if not os.path.exists(self.filename):
                wb = Workbook()
                wb.save(self.filename)
                print(f"✅ Создан новый Excel файл: {self.filename}")
                # Создаем папку на Яндекс.Диске при первом запуске
                if yandex_disk:
                    yandex_disk.create_folder(YANDEX_DISK_FOLDER)
            else:
                print(f"📁 Excel файл уже существует: {self.filename}")

            if os.path.exists(self.filename):
                file_stats = os.stat(self.filename)
                print(f"📊 Размер файла: {file_stats.st_size} байт")
        except Exception as e:
            print(f"❌ Ошибка при создании файла: {e}")
            import traceback
            traceback.print_exc()

    def get_user_sheet(self, user_id: int, last_name: str = ""):
        """Возвращает или создаёт лист для пользователя"""
        try:
            wb = openpyxl.load_workbook(self.filename)
        except Exception as e:
            print(f"Ошибка загрузки файла: {e}")
            self._ensure_file_exists()
            wb = openpyxl.load_workbook(self.filename)

        if last_name and last_name.strip():
            sheet_name = ''.join(c for c in last_name.strip() if c.isalnum() or c in ' _-')[:31]
            if not sheet_name:
                sheet_name = f"user_{user_id}"
        else:
            sheet_name = f"user_{user_id}"

        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(sheet_name)
            sheet['A1'] = "Дата"
            sheet['B1'] = "Время работы"
            sheet['C1'] = "Описание работы"
            sheet['D1'] = "Часы работы без обеда"
            sheet.column_dimensions['A'].width = 12
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 50
            sheet.column_dimensions['D'].width = 20
            bold_font = openpyxl.styles.Font(bold=True)
            for cell in ['A1', 'B1', 'C1', 'D1']:
                sheet[cell].font = bold_font
            print(f"✅ Создан новый лист: {sheet_name}")
        wb.save(self.filename)
        return sheet_name

    def calculate_work_hours(self, time_range: str, had_lunch: bool = False):
        """Поддерживает несколько периодов, разделённых запятыми."""
        try:
            total_seconds = 0
            periods = re.split(r',\s*', time_range.strip())
            for period in periods:
                if not period:
                    continue
                clean_period = re.sub(r'[с\-\–\—]', ' ', period).strip()
                times = re.findall(r'(\d{1,2}:\d{2}|\d{1,2})', clean_period)
                if len(times) >= 2:
                    start_str = times[0]
                    end_str = times[1]
                    if ':' not in start_str:
                        start_str += ':00'
                    if ':' not in end_str:
                        end_str += ':00'
                    start = datetime.strptime(start_str, '%H:%M')
                    end = datetime.strptime(end_str, '%H:%M')
                    if end < start:
                        end += timedelta(days=1)
                    total_seconds += (end - start).total_seconds()

            total_hours = total_seconds / 3600
            work_hours = total_hours - (0.5 if had_lunch else 0)
            return round(max(work_hours, 0), 2)
        except Exception as e:
            print(f"Ошибка вычисления часов: {e}")
            return 0.0

    def has_today_entry(self, user_id: int, last_name: str = ""):
        """Проверяет, есть ли уже запись за сегодня"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet_name = self.get_user_sheet(user_id, last_name)
            sheet = wb[sheet_name]
            
            current_date = datetime.now().strftime("%d.%m.%Y")
            
            for row in range(2, sheet.max_row + 1):
                date_cell = sheet[f'A{row}']
                if date_cell.value == current_date:
                    return True
            return False
        except Exception as e:
            print(f"❌ Ошибка при проверке записи за сегодня: {e}")
            return False

    def add_entry(self, user_id: int, time_range: str, description: str, had_lunch: bool, last_name: str = ""):
        try:
            print(f"🔧 Попытка сохранить запись для user_id: {user_id}")
            print(f"📁 Путь к файлу: {self.filename}")
            print(f"📝 Данные: {time_range}, {description}, обед: {had_lunch}")

            # Проверяем лимит записей
            if self.has_today_entry(user_id, last_name):
                return False, "limit_exceeded"

            # Гарантируем существование листа
            sheet_name = self.get_user_sheet(user_id, last_name)
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb[sheet_name]

            row = sheet.max_row + 1
            work_hours = self.calculate_work_hours(time_range, had_lunch)
            current_date = datetime.now().strftime("%d.%m.%Y")
            sheet[f'A{row}'] = current_date
            sheet[f'B{row}'] = time_range
            sheet[f'C{row}'] = description
            sheet[f'D{row}'] = work_hours
            wb.save(self.filename)
            
            # ✅ Сохраняем на Яндекс.Диск после добавления записи
            if yandex_disk:
                remote_file_path = f"{YANDEX_DISK_FOLDER}/work_tracker_backup.xlsx"
                if yandex_disk.upload_file(self.filename, remote_file_path):
                    print(f"✅ Резервная копия загружена на Яндекс.Диск")
                else:
                    print(f"⚠️ Не удалось загрузить резервную копию на Яндекс.Диск")
            
            print(f"✅ Запись добавлена для пользователя {user_id}: {work_hours:.2f} ч.")
            return True, "success"
        except Exception as e:
            print(f"❌ Ошибка при записи в Excel: {e}")
            import traceback
            traceback.print_exc()
            return False, "error"

    def delete_today_entry(self, user_id: int, last_name: str = ""):
        """Удаляет последнюю запись за сегодня"""
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet_name = self.get_user_sheet(user_id, last_name)
            sheet = wb[sheet_name]
            
            current_date = datetime.now().strftime("%d.%m.%Y")
            deleted_data = None
            
            for row in range(sheet.max_row, 1, -1):
                date_cell = sheet[f'A{row}']
                if date_cell.value == current_date:
                    deleted_data = {
                        'date': sheet[f'A{row}'].value,
                        'time_range': sheet[f'B{row}'].value,
                        'description': sheet[f'C{row}'].value,
                        'work_hours': sheet[f'D{row}'].value
                    }
                    sheet.delete_rows(row)
                    wb.save(self.filename)
                    
                    # ✅ Сохраняем на Яндекс.Диск после удаления записи
                    if yandex_disk:
                        remote_file_path = f"{YANDEX_DISK_FOLDER}/work_tracker_backup.xlsx"
                        if yandex_disk.upload_file(self.filename, remote_file_path):
                            print(f"✅ Резервная копия загружена на Яндекс.Диск после удаления")
                    
                    print(f"✅ Запись за сегодня удалена для пользователя {user_id}")
                    return True, deleted_data
            
            return False, None
        except Exception as e:
            print(f"❌ Ошибка при удалении записи: {e}")
            return False, None

    def get_user_stats(self, user_id: int, last_name: str = ""):
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet_name = self.get_user_sheet(user_id, last_name)
            sheet = wb[sheet_name]
            return sheet.max_row - 1
        except Exception as e:
            print(f"❌ Ошибка при получении статистики: {e}")
            return 0

excel_manager = ExcelManager(EXCEL_FILE)
user_data_cache = {}

def get_main_menu_keyboard():
    keyboard = [
        ["📝 Отчет"],
        ["🗑️ Удалить запись", "⚙️ Напоминание"],
        ["📥 Скачать отчет", "☁️ Синхронизировать"]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, input_field_placeholder="Выберите действие...")

def get_yes_no_keyboard():
    return ReplyKeyboardMarkup([["Да", "Нет"]], resize_keyboard=True, one_time_keyboard=True)

async def send_welcome_message(update: Update, user):
    yandex_status = "✅ ВКЛЮЧЕН" if yandex_disk else "❌ ВЫКЛЮЧЕН"
    welcome_text = (
        "🎉 *ДОБРО ПОЖАЛОВАТЬ!* 🎉\n"
        "🤖 *Я - Work Tracker Bot* 🤖\n"
        "*Моя задача:* Помогать тебе вести учет рабочего времени!\n"
        "*Как это работает:*\n"
        "• Каждый день я буду напоминать тебе заполнить отчет\n"
        "• Ты указываешь, в какое время работал и что делал\n"
        "• Все данные автоматически сохраняются в Excel таблицу\n"
        "• У каждого сотрудника свой лист в таблице\n"
        f"• ☁️ *Резервное копирование:* {yandex_status}\n"
        "*Важно:* Можно сделать только *1 запись в день*\n"
        "*Преимущества:*\n"
        "✅ Всегда актуальная информация о работе\n"
        "✅ Удобный учет времени\n"
        "✅ Автоматическое сохранение\n"
        "✅ Индивидуальные настройки\n"
        "✅ Резервное копирование на Яндекс.Диск\n"
        "Используй кнопки меню ниже для навигации!"
    )
    await update.message.reply_text(welcome_text, parse_mode='Markdown', reply_markup=get_main_menu_keyboard())

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    user_id = user.id
    is_new_user = user_id not in WELCOMED_USERS
    if is_new_user:
        await send_welcome_message(update, user)
        WELCOMED_USERS.add(user_id)
        await asyncio.sleep(2)
    if user_id not in USER_SETTINGS:
        USER_SETTINGS[user_id] = {
            'reminder_time': time(hour=DEFAULT_REMINDER_HOUR, minute=DEFAULT_REMINDER_MINUTE),
            'username': user.username or "",
            'first_name': user.first_name or "",
            'last_name': user.last_name or "",
            'first_seen': datetime.now()
        }
    last_name = user.last_name or user.first_name or ""
    stats = excel_manager.get_user_stats(user_id, last_name)
    reminder_time = USER_SETTINGS[user_id]['reminder_time']
    has_today_entry = excel_manager.has_today_entry(user_id, last_name)
    
    if is_new_user:
        message_text = f"👋 *Рад познакомиться, {user.first_name}!*\n"
    else:
        message_text = f"👋 *С возвращением, {user.first_name}!*\n"
    
    message_text += (
        f"📊 Твоя статистика: *{stats} записей*\n"
        f"⏰ Напоминание установлено на: *{reminder_time.strftime('%H:%M')}*\n"
    )
    
    if has_today_entry:
        message_text += f"📝 *Сегодняшняя запись:* ✅ УЖЕ СДЕЛАНА\n"
    else:
        message_text += f"📝 *Сегодняшняя запись:* ❌ ЕЩЕ НЕТ\n"
        
    yandex_status = "✅ ВКЛЮЧЕНО" if yandex_disk else "❌ ВЫКЛЮЧЕНО"
    message_text += f"☁️ *Резервное копирование:* {yandex_status}\n\n"
        
    message_text += (
        f"*Используй кнопки меню для управления:*\n"
        f"📝 *Отчет* - добавить запись о работе\n"
        f"🗑️ *Удалить запись* - удалить сегодняшнюю запись\n"
        f"⚙️ *Напоминание* - изменить время напоминания\n"
        f"📥 *Скачать отчет* - получить Excel файл\n"
        f"☁️ *Синхронизировать* - принудительно сохранить на Яндекс.Диск"
    )
    await update.message.reply_text(message_text, parse_mode='Markdown', reply_markup=get_main_menu_keyboard())

async def handle_menu_buttons(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "📝 Отчет":
        return await report_command(update, context)
    elif text == "🗑️ Удалить запись":
        return await delete_entry_command(update, context)
    elif text == "⚙️ Напоминание":
        return await reminder_command(update, context)
    elif text == "📥 Скачать отчет":
        return await download_file(update, context)
    elif text == "☁️ Синхронизировать":
        return await sync_to_yandex_disk(update, context)
    else:
        await update.message.reply_text("Неизвестная команда. Используй кнопки меню.", reply_markup=get_main_menu_keyboard())

async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user = update.message.from_user
    last_name = user.last_name or user.first_name or ""
    
    # Проверяем, есть ли уже запись за сегодня
    if excel_manager.has_today_entry(user_id, last_name):
        await update.message.reply_text(
            "❌ *Вы уже сделали запись за сегодняшний день.*\n\n"
            "Чтобы создать новую запись, сначала удалите предыдущую через кнопку \"🗑️ Удалить запись\", "
            "а затем создайте новую через кнопку \"📝 Отчет\".",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
        return ConversationHandler.END
    
    await update.message.reply_text(
        "📝 *Заполним отчет о работе!*\n"
        "🕐 *ШАГ 1:* Укажи ВРЕМЯ РАБОТЫ (можно несколько периодов):\n"
        "*Примеры:*\n"
        "• 9:00-18:00\n"
        "• 9:00-14:00, 15:00-18:00\n"
        "• с 10 до 12, 14:00-17:30\n"
        "Используй запятую для разделения периодов.\n"
        "*Примечание:* После ввода я уточню, был ли у тебя обед.",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()
    )
    return WAITING_TIME

async def receive_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    time_range = update.message.text
    if user_id not in user_data_cache:
        user_data_cache[user_id] = {}
    user_data_cache[user_id]['time_range'] = time_range

    total_hours = excel_manager.calculate_work_hours(time_range, had_lunch=False)
    await update.message.reply_text(
        f"✅ *Отлично!*\n"
        f"⏱️ *Общее время работы:* {total_hours:.2f} ч.\n"
        "🍽️ *Был ли у тебя сегодня обед?*\n"
        "(Обед = вычет 0.5 часа)",
        reply_markup=get_yes_no_keyboard()
    )
    return WAITING_LUNCH_CONFIRMATION

async def receive_lunch_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    text = update.message.text.strip().lower()
    if text in ["да", "yes", "д"]:
        had_lunch = True
    elif text in ["нет", "no", "н"]:
        had_lunch = False
    else:
        await update.message.reply_text("Пожалуйста, выбери «Да» или «Нет».", reply_markup=get_yes_no_keyboard())
        return WAITING_LUNCH_CONFIRMATION

    if user_id not in user_data_cache:
        user_data_cache[user_id] = {}
    user_data_cache[user_id]['had_lunch'] = had_lunch

    await update.message.reply_text(
        "📝 *ШАГ 2:* Теперь опиши ОПИСАНИЕ РАБОТЫ — что ты делал:\n"
        "*Примеры:*\n"
        "• Разрабатывал новый функционал\n"
        "• Участвовал в совещаниях\n"
        "• Изучал документацию\n"
        "• Исправлял ошибки\n"
        "• Общался с клиентами",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()
    )
    return WAITING_DESCRIPTION

async def receive_description(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    description = update.message.text
    user = update.message.from_user
    if (user_id not in user_data_cache or
        'time_range' not in user_data_cache[user_id] or
        'had_lunch' not in user_data_cache[user_id]):
        await update.message.reply_text("❌ Что-то пошло не так. Давай начнем заново", reply_markup=get_main_menu_keyboard())
        return ConversationHandler.END

    time_range = user_data_cache[user_id]['time_range']
    had_lunch = user_data_cache[user_id]['had_lunch']
    last_name = user.last_name or user.first_name or ""

    success, result = excel_manager.add_entry(user_id, time_range, description, had_lunch, last_name)
    
    if result == "limit_exceeded":
        await update.message.reply_text(
            "❌ *Вы уже сделали запись за сегодняшний день.*\n\n"
            "Чтобы создать новую запись, сначала удалите предыдущую через кнопку \"🗑️ Удалить запись\", "
            "а затем создайте новую через кнопку \"📝 Отчет\".",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
    elif success:
        stats = excel_manager.get_user_stats(user_id, last_name)
        current_date = datetime.now().strftime("%d.%m.%Y")
        work_hours = excel_manager.calculate_work_hours(time_range, had_lunch)
        
        yandex_sync_text = ""
        if yandex_disk:
            yandex_sync_text = "☁️ *Данные автоматически сохранены на Яндекс.Диск*\n"
        
        await update.message.reply_text(
            "🎉 *ОТЛИЧНО! Запись сохранена!*\n"
            f"{yandex_sync_text}\n"
            f"📅 *Дата:* {current_date}\n"
            f"🕐 *Время работы:* {time_range}\n"
            f"🍽️ *Обед:* {'Да' if had_lunch else 'Нет'}\n"
            f"⏱️ *Часы работы без обеда:* {work_hours:.2f} ч.\n"
            f"📝 *Описание работы:* {description}\n"
            f"📊 *Всего записей:* {stats}\n\n"
            "*Теперь ты можешь:*\n"
            "• 🗑️ *Удалить запись* - если нужно исправить\n"
            "• 📥 *Скачать отчет* - получить полный файл\n"
            "• ☁️ *Синхронизировать* - принудительно сохранить в облако\n"
            "*Новая запись будет доступна завтра*",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
    else:
        await update.message.reply_text(
            "❌ Произошла ошибка при сохранении. Попробуй еще раз",
            reply_markup=get_main_menu_keyboard()
        )
    
    if user_id in user_data_cache:
        del user_data_cache[user_id]
    return ConversationHandler.END

async def delete_entry_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user = update.message.from_user
    last_name = user.last_name or user.first_name or ""
    
    success, deleted_data = excel_manager.delete_today_entry(user_id, last_name)
    
    if success:
        yandex_sync_text = ""
        if yandex_disk:
            yandex_sync_text = "\n☁️ *Изменения сохранены на Яндекс.Диск*"
            
        await update.message.reply_text(
            "🗑️ *Запись за сегодня успешно удалена!*\n"
            f"{yandex_sync_text}\n\n"
            f"📅 *Дата:* {deleted_data['date']}\n"
            f"🕐 *Время работы:* {deleted_data['time_range']}\n"
            f"📝 *Описание:* {deleted_data['description']}\n"
            f"⏱️ *Часы работы:* {deleted_data['work_hours']} ч.\n\n"
            "Теперь ты можешь создать новую запись через кнопку \"📝 Отчет\"",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
    else:
        await update.message.reply_text(
            "❌ *Не найдено записей за сегодня для удаления.*\n\n"
            "Сначала создайте запись через кнопку \"📝 Отчет\"",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )

async def sync_to_yandex_disk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Принудительная синхронизация с Яндекс.Диском"""
    if not yandex_disk:
        await update.message.reply_text(
            "❌ *Синхронизация с Яндекс.Диском отключена.*\n\n"
            "Для включения:\n"
            "1. Получите OAuth-токен Яндекс.Диск\n"
            "2. Установите переменную YANDEX_DISK_TOKEN\n"
            "3. Перезапустите бота",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
        return
    
    await update.message.reply_text(
        "☁️ *Начинаю синхронизацию с Яндекс.Диском...*",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()
    )
    
    try:
        remote_file_path = f"{YANDEX_DISK_FOLDER}/work_tracker_backup.xlsx"
        
        if yandex_disk.upload_file(EXCEL_FILE, remote_file_path):
            file_info = yandex_disk.get_file_info(remote_file_path)
            if file_info:
                file_size = file_info.get('size', 0)
                modified = file_info.get('modified', '')
                await update.message.reply_text(
                    f"✅ *Синхронизация успешно завершена!*\n\n"
                    f"📊 *Данные файла на Яндекс.Диске:*\n"
                    f"• 📁 Размер: {int(file_size) / 1024 / 1024:.2f} MB\n"
                    f"• 📅 Обновлен: {modified[:19] if modified else 'Неизвестно'}\n"
                    f"• 🔗 Путь: {YANDEX_DISK_FOLDER}/work_tracker_backup.xlsx\n\n"
                    f"Все данные надежно сохранены в облаке! ☁️",
                    parse_mode='Markdown',
                    reply_markup=get_main_menu_keyboard()
                )
            else:
                await update.message.reply_text(
                    "✅ *Файл загружен на Яндекс.Диск!*\n\n"
                    "Резервная копия успешно сохранена в облаке. ☁️",
                    parse_mode='Markdown',
                    reply_markup=get_main_menu_keyboard()
                )
        else:
            await update.message.reply_text(
                "❌ *Ошибка синхронизации!*\n\n"
                "Не удалось загрузить файл на Яндекс.Диск. "
                "Проверьте настройки и попробуйте позже.",
                parse_mode='Markdown',
                reply_markup=get_main_menu_keyboard()
            )
            
    except Exception as e:
        print(f"❌ Ошибка при синхронизации: {e}")
        await update.message.reply_text(
            "❌ *Произошла ошибка при синхронизации!*\n\n"
            "Попробуйте позже или проверьте настройки Яндекс.Диска.",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id in user_data_cache:
        del user_data_cache[user_id]
    await update.message.reply_text("❌ Диалог отменен.", reply_markup=get_main_menu_keyboard())
    return ConversationHandler.END

async def reminder_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "⏰ *Установи свое индивидуальное время напоминания!*\n"
        "Введи время в формате *ЧАСЫ:МИНУТЫ* (24-часовой формат):\n"
        "*Примеры:*\n"
        "• 18:00 - в 6 вечера\n"
        "• 09:30 - в 9:30 утра\n"
        "• 17:45 - в 5:45 вечера\n"
        "*Введи время:*",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()
    )
    return WAITING_REMINDER_TIME

async def receive_reminder_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user_input = update.message.text.strip()
    time_pattern = r'^([0-1]?[0-9]|2[0-3]):([0-5][0-9])$'
    if not re.match(time_pattern, user_input):
        await update.message.reply_text(
            "❌ *Неверный формат времени!*\n"
            "Пожалуйста, введи время в формате *ЧАСЫ:МИНУТЫ* (24-часовой формат):\n"
            "• 18:00\n• 09:30\n• 17:45\nПопробуй еще раз:",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
        return ConversationHandler.END
    hours, minutes = map(int, user_input.split(':'))
    if user_id not in USER_SETTINGS:
        USER_SETTINGS[user_id] = {}
    reminder_time = time(hour=hours, minute=minutes)
    USER_SETTINGS[user_id]['reminder_time'] = reminder_time
    USER_SETTINGS[user_id]['first_name'] = update.message.from_user.first_name or ""
    USER_SETTINGS[user_id]['last_name'] = update.message.from_user.last_name or ""

    global global_app
    job_queue = global_app.job_queue
    if job_queue:
        for job in job_queue.get_jobs_by_name(str(user_id)):
            job.schedule_removal()
        job_time = time(hour=hours, minute=minutes, tzinfo=TIMEZONE)
        job_queue.run_daily(
            send_daily_reminder,
            time=job_time,
            days=tuple(range(7)),
            data=user_id,
            name=str(user_id)
        )
        job_queue.run_once(
            send_test_reminder,
            when=60,
            data=user_id,
            name=f"test_{user_id}"
        )
        print(f"✅ Напоминание установлено для {user_id} на {hours:02d}:{minutes:02d}")
    else:
        print("❌ job_queue недоступен — критическая ошибка!")

    await update.message.reply_text(
        f"✅ *Отлично! Твое время напоминания установлено на {user_input}*\n"
        f"Каждый день в это время я буду присылать тебе напоминание заполнить отчет о работе.\n"
        f"*Тестовое напоминание придет через 1 минуту* ⏰\n"
        f"Ты всегда можешь изменить время через кнопку '⚙️ Напоминание'",
        parse_mode='Markdown',
        reply_markup=get_main_menu_keyboard()
    )
    return ConversationHandler.END

async def send_test_reminder(context):
    try:
        user_id = context.job.data
        await context.bot.send_message(
            chat_id=user_id,
            text="🧪 *ТЕСТОВОЕ НАПОМИНАНИЕ!*\n"
                 "Это тестовое сообщение чтобы проверить работу напоминаний.\n"
                 "Если ты видишь это сообщение - значит система напоминаний работает правильно! ✅",
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
        print(f"✅ Тестовое напоминание отправлено пользователю {user_id}")
    except Exception as e:
        print(f"❌ Ошибка при отправке тестового напоминания: {e}")

async def send_daily_reminder(context):
    try:
        user_id = context.job.data
        reminder_time_str = "18:00"
        if user_id in USER_SETTINGS and 'reminder_time' in USER_SETTINGS[user_id]:
            reminder_time_str = USER_SETTINGS[user_id]['reminder_time'].strftime('%H:%M')
        
        user = USER_SETTINGS.get(user_id, {})
        last_name = user.get('last_name', '') or user.get('first_name', '')
        has_today_entry = excel_manager.has_today_entry(user_id, last_name)
        
        if has_today_entry:
            message_text = (
                f"🕔 *ЕЖЕДНЕВНОЕ НАПОМИНАНИЕ ({reminder_time_str})!*\n"
                f"Привет! Я вижу, что ты уже заполнил отчет за сегодня. ✅\n\n"
                f"Если нужно что-то исправить:\n"
                f"1️⃣ Нажми '🗑️ Удалить запись'\n"
                f"2️⃣ Затем создай новую через '📝 Отчет'"
            )
        else:
            message_text = (
                f"🕔 *ЕЖЕДНЕВНОЕ НАПОМИНАНИЕ ({reminder_time_str})!*\n"
                f"Привет! Пора заполнить отчет о работе за сегодня.\n"
                f"Нажми кнопку '📝 Отчет' чтобы указать:\n"
                f"1️⃣ В какое время ты работал (можно несколько периодов)\n"
                f"2️⃣ Был ли обед\n"
                f"3️⃣ Что ты делал\n"
                f"Это займет всего 30 секунд! ⏱️"
            )
            
        await context.bot.send_message(
            chat_id=user_id,
            text=message_text,
            parse_mode='Markdown',
            reply_markup=get_main_menu_keyboard()
        )
        print(f"✅ Ежедневное напоминание отправлено пользователю {user_id}")
    except Exception as e:
        print(f"❌ Ошибка при отправке напоминания пользователю {user_id}: {e}")

async def download_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not os.path.exists(EXCEL_FILE):
            await update.message.reply_text(
                "❌ Файл с отчетами еще не создан. Добавь первую запись через кнопку '📝 Отчет'",
                reply_markup=get_main_menu_keyboard()
            )
            return
        
        yandex_status = ""
        if yandex_disk:
            yandex_status = "\n☁️ *Резервная копия хранится на Яндекс.Диске*"
            
        with open(EXCEL_FILE, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=f"work_reports_{datetime.now().strftime('%d.%m.%Y')}.xlsx",
                caption=f"📊 *Вот твой файл с отчетами!*\n"
                       f"Файл содержит все записи о рабочем времени.\n"
                       f"Каждый пользователь имеет свой лист в файле.\n"
                       f"*Ограничение:* 1 запись в день на пользователя"
                       f"{yandex_status}",
                parse_mode='Markdown',
                reply_markup=get_main_menu_keyboard()
            )
        print(f"✅ Файл отправлен пользователю {update.message.from_user.id}")
    except Exception as e:
        print(f"❌ Ошибка при отправке файла: {e}")
        await update.message.reply_text(
            "❌ Произошла ошибка при отправке файла. Попробуй позже.",
            reply_markup=get_main_menu_keyboard()
        )

async def handle_unknown_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❌ *Неизвестная команда.*\n"
        "*Используй кнопки меню:*\n"
        "📝 Отчет - добавить запись о работе\n"
        "🗑️ Удалить запись - удалить сегодняшнюю запись\n"
        "⚙️ Напоминание - изменить время напоминания\n"
        "📥 Скачать отчет - получить Excel файл\n"
        "☁️ Синхронизировать - принудительно сохранить на Яндекс.Диск",
        parse_mode='Markdown',
        reply_markup=get_main_menu_keyboard()
    )

def restore_reminders(application: Application):
    job_queue = application.job_queue
    restored_count = 0
    for user_id, settings in USER_SETTINGS.items():
        if 'reminder_time' in settings:
            for job in job_queue.get_jobs_by_name(str(user_id)):
                job.schedule_removal()
            job_time = time(
                hour=settings['reminder_time'].hour,
                minute=settings['reminder_time'].minute,
                tzinfo=TIMEZONE
            )
            job_queue.run_daily(
                send_daily_reminder,
                time=job_time,
                days=tuple(range(7)),
                data=user_id,
                name=str(user_id)
            )
            restored_count += 1
            print(f"🔁 Восстановлено напоминание для {user_id} на {settings['reminder_time'].strftime('%H:%M')}")
    print(f"✅ Восстановлено {restored_count} напоминаний.")

def main():
    global global_app
    print("🚀 Запуск Work Tracker Bot...")
    print("📊 Бот для учета рабочего времени")
    print("💾 Excel файл:", EXCEL_FILE)
    print("⏱️ Поддержка нескольких периодов + выбор обеда")
    print("📝 Ограничение: 1 запись в день на пользователя")
    print(f"☁️  Яндекс.Диск: {'ВКЛЮЧЕН' if yandex_disk else 'ВЫКЛЮЧЕН'}")

    application = Application.builder().token(BOT_TOKEN).build()
    global_app = application

    report_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("report", report_command),
            MessageHandler(filters.Regex("^(📝 Отчет)$"), report_command)
        ],
        states={
            WAITING_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_time)],
            WAITING_LUNCH_CONFIRMATION: [MessageHandler(filters.Regex("^(Да|Нет)$"), receive_lunch_confirmation)],
            WAITING_DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_description)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )

    reminder_conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("reminder", reminder_command),
            MessageHandler(filters.Regex("^(⚙️ Напоминание)$"), reminder_command)
        ],
        states={
            WAITING_REMINDER_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_reminder_time)],
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("download", download_file))
    application.add_handler(CommandHandler("delete", delete_entry_command))
    application.add_handler(CommandHandler("sync", sync_to_yandex_disk))
    application.add_handler(MessageHandler(filters.Regex("^(🗑️ Удалить запись)$"), delete_entry_command))
    application.add_handler(MessageHandler(filters.Regex("^(📥 Скачать отчет)$"), download_file))
    application.add_handler(MessageHandler(filters.Regex("^(☁️ Синхронизировать)$"), sync_to_yandex_disk))
    application.add_handler(report_conv_handler)
    application.add_handler(reminder_conv_handler)
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu_buttons))
    application.add_handler(MessageHandler(filters.COMMAND, handle_unknown_command))

    restore_reminders(application)

    print("✅ Бот успешно запущен!")
    print("📱 Ожидаем сообщения от пользователей...")
    try:
        application.run_polling()
    except KeyboardInterrupt:
        print("\n❌ Бот остановлен")
    except Exception as e:
        print(f"❌ Ошибка: {e}")

if __name__ == "__main__":
    main()
